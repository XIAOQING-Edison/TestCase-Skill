"""
Excel 输出格式化器
生成 Excel 格式的测试用例文件
"""

import json
from datetime import datetime
from typing import Dict, List, Any


class ExcelFormatter:
    """Excel 格式化器"""

    def __init__(self, module_name: str = "TestModule"):
        self.module_name = module_name
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def format(self, testcases: List[Dict], filepath: str = None):
        """
        格式化测试用例为 Excel

        Args:
            testcases: 测试用例列表
            filepath: 输出文件路径（可选）

        Returns:
            文件路径
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # 返回 JSON 格式作为降级
            return json.dumps(testcases, ensure_ascii=False, indent=2)

        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 确定用例类型并设置表头
        if testcases and '用例编号' in testcases[0]:
            headers = [
                "用例编号", "用例标题", "模块/优先级", "接口名称", "前置条件",
                "请求URL", "请求类型", "请求头", "请求参数类型", "请求数据",
                "预期响应状态", "预期响应信息", "实际响应信息", "是否通过",
                "测试类型", "断言来源", "是否为假设", "备注"
            ]
            fill_color = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        else:
            headers = [
                "功能ID", "用例标题", "功能模块", "优先级", "预置条件",
                "测试数据", "执行步骤", "预期结果", "测试结果",
                "测试版本号", "测试人员", "断言来源", "是否为假设", "备注", "测试类型"
            ]
            fill_color = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = fill_color
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据
        for row, tc in enumerate(testcases, 2):
            for col, header in enumerate(headers, 1):
                key = header.replace(' ', '').replace('/', '')
                # 尝试多种可能的关键字
                value = tc.get(header, '')
                if not value:
                    value = tc.get(key, '')
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 设置列宽
        col_widths = [15, 40, 15, 25, 20, 40, 10, 30, 15, 35, 15, 35, 35, 10, 12, 24, 12, 20, 20]
        for i, width in enumerate(col_widths[:len(headers)], 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 如果用例太多，创建统计工作表
        if len(testcases) > 100:
            self._create_summary_sheet(wb, testcases, headers)

        # 保存文件
        if filepath is None:
            import os
            filepath = f"测试用例_{self.module_name}_{self.timestamp}.xlsx"
        wb.save(filepath)
        return filepath

    def _create_summary_sheet(self, wb, testcases: List[Dict], headers: List[str]):
        """创建统计工作表"""
        ws_summary = wb.create_sheet("统计汇总")

        # 统计各优先级数量
        priority_count = {}
        for tc in testcases:
            priority = tc.get('模块/优先级', tc.get('优先级', '未知'))
            priority_count[priority] = priority_count.get(priority, 0) + 1

        ws_summary.cell(row=1, column=1, value="优先级统计")
        ws_summary.cell(row=2, column=1, value="优先级")
        ws_summary.cell(row=2, column=2, value="数量")
        ws_summary.cell(row=2, column=3, value="占比")

        total = len(testcases)
        for idx, (priority, count) in enumerate(priority_count.items(), 3):
            ws_summary.cell(row=idx, column=1, value=priority)
            ws_summary.cell(row=idx, column=2, value=count)
            ws_summary.cell(row=idx, column=3, value=f"{count/total*100:.1f}%")

    def format_json(self, testcases: List[Dict]) -> str:
        """格式化测试用例为 JSON（降级方案）"""
        return json.dumps(testcases, ensure_ascii=False, indent=2)


# 便捷函数
def format_excel(testcases: List[Dict], module_name: str, filepath: str = None):
    """格式化测试用例为 Excel"""
    return ExcelFormatter(module_name).format(testcases, filepath)
