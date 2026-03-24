#!/usr/bin/env python3
"""
测试用例生成器 - 兼容入口
支持从各种文档格式生成测试用例
"""

import json
import os
import sys
from datetime import datetime

# 添加当前目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from generators.api_generator import generate_api_testcases
from generators.functional_generator import generate_functional_testcases


def create_functional_testcase_excel(testcases, module_name):
    """生成功能测试用例Excel（兼容旧版本）"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return json.dumps(testcases, ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "功能测试用例"

    headers = [
        "功能ID", "用例标题", "功能模块", "优先级", "预置条件",
        "测试数据", "执行步骤", "预期结果", "测试结果",
        "测试版本号", "测试人员", "备注"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row, tc in enumerate(testcases, 2):
        ws.cell(row=row, column=1, value=tc.get("功能ID", "")).border = thin_border
        ws.cell(row=row, column=2, value=tc.get("用例标题", "")).border = thin_border
        ws.cell(row=row, column=3, value=tc.get("功能模块", module_name)).border = thin_border
        ws.cell(row=row, column=4, value=tc.get("优先级", "")).border = thin_border
        ws.cell(row=row, column=5, value=tc.get("预置条件", "")).border = thin_border
        ws.cell(row=row, column=6, value=tc.get("测试数据", "")).border = thin_border
        ws.cell(row=row, column=7, value=tc.get("执行步骤", "")).border = thin_border
        ws.cell(row=row, column=8, value=tc.get("预期结果", "")).border = thin_border
        ws.cell(row=row, column=9, value=tc.get("测试结果", "")).border = thin_border
        ws.cell(row=row, column=10, value=tc.get("测试版本号", "")).border = thin_border
        ws.cell(row=row, column=11, value=tc.get("测试人员", "")).border = thin_border
        ws.cell(row=row, column=12, value=tc.get("备注", "")).border = thin_border

    col_widths = [15, 30, 15, 8, 25, 25, 35, 35, 10, 12, 10, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"功能测试用例_{module_name}_{timestamp}.xlsx"
    wb.save(filename)
    return filename


def create_api_testcase_excel(testcases, module_name):
    """生成接口测试用例Excel（兼容旧版本）"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return json.dumps(testcases, ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "接口测试用例"

    headers = [
        "用例编号", "用例标题", "模块/优先级", "接口名称", "前置条件",
        "请求URL", "请求类型", "请求头", "请求参数类型", "请求数据",
        "预期响应信息", "实际响应信息", "是否通过"
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for row, tc in enumerate(testcases, 2):
        ws.cell(row=row, column=1, value=tc.get("用例编号", "")).border = thin_border
        ws.cell(row=row, column=2, value=tc.get("用例标题", "")).border = thin_border
        ws.cell(row=row, column=3, value=tc.get("模块/优先级", "")).border = thin_border
        ws.cell(row=row, column=4, value=tc.get("接口名称", "")).border = thin_border
        ws.cell(row=row, column=5, value=tc.get("前置条件", "")).border = thin_border
        ws.cell(row=row, column=6, value=tc.get("请求URL", "")).border = thin_border
        ws.cell(row=row, column=7, value=tc.get("请求类型", "")).border = thin_border
        ws.cell(row=row, column=8, value=tc.get("请求头", "")).border = thin_border
        ws.cell(row=row, column=9, value=tc.get("请求参数类型", "")).border = thin_border
        ws.cell(row=row, column=10, value=tc.get("请求数据", "")).border = thin_border
        ws.cell(row=row, column=11, value=tc.get("预期响应信息", "")).border = thin_border
        ws.cell(row=row, column=12, value=tc.get("实际响应信息", "")).border = thin_border
        ws.cell(row=row, column=13, value=tc.get("是否通过", "")).border = thin_border

    col_widths = [15, 30, 15, 15, 20, 30, 10, 25, 15, 30, 35, 35, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"接口测试用例_{module_name}_{timestamp}.xlsx"
    wb.save(filename)
    return filename


def generate_testcases_from_requirement(requirement_text: str, module_name: str):
    """根据需求文本生成测试用例"""
    # 简化的功能用例生成
    testcases = {
        "functional": [],
        "api": []
    }

    # 解析需求中的功能点
    features = parse_features(requirement_text, module_name)

    for feature in features:
        testcases["functional"].extend(generate_functional_testcases(feature, module_name))

    return testcases


def parse_features(text: str, module_name: str):
    """解析需求文本，提取功能点"""
    features = []
    lines = text.split('\n')

    current_feature = {}
    for line in lines:
        line = line.strip()
        if not line:
            continue

        if line.startswith('#') or '功能' in line or '模块' in line:
            if current_feature:
                features.append(current_feature)
            current_feature = {"name": line.lstrip('#').strip(), "description": ""}
        elif current_feature:
            current_feature["description"] += line + "\n"

    if current_feature:
        features.append(current_feature)

    return features if features else [{"name": module_name, "description": text[:500]}]


def generate_functional_testcases(feature: dict, module_name: str):
    """生成功能测试用例"""
    testcases = []
    feature_name = feature.get("name", module_name)
    testcase_id = 1

    positive_tests = [
        {"title": f"验证{feature_name}主流程正常执行", "priority": "P0", "description": "使用有效数据，验证核心功能正常"},
        {"title": f"验证{feature_name}所有必填字段正常处理", "priority": "P0", "description": "填写所有必填项，验证功能正常"},
        {"title": f"验证{feature_name}可选字段的处理", "priority": "P1", "description": "填写必填项和部分可选字段，验证功能正常"}
    ]

    negative_tests = [
        {"title": f"验证{feature_name}缺少必填字段时的处理", "priority": "P0", "description": "不填写必填项，验证系统正确提示"},
        {"title": f"验证{feature_name}输入非法数据的处理", "priority": "P1", "description": "输入特殊字符、超长字符等，验证系统处理"},
        {"title": f"验证{feature_name}边界条件处理", "priority": "P1", "description": "测试边界值、极限值情况"},
        {"title": f"验证{feature_name}重复操作的处理", "priority": "P2", "description": "重复执行操作，验证系统响应"}
    ]

    for test in positive_tests + negative_tests:
        testcases.append({
            "功能ID": f"FUNC_{module_name.upper()}_{testcase_id:03d}",
            "用例标题": test["title"],
            "功能模块": feature_name,
            "优先级": test["priority"],
            "预置条件": "用户已登录系统",
            "测试数据": "根据具体需求确定",
            "执行步骤": f"1. 进入{feature_name}页面\n2. 输入有效数据\n3. 提交请求\n4. 验证结果",
            "预期结果": test["description"],
            "测试结果": "待执行",
            "测试版本号": "v1.0.0",
            "测试人员": "",
            "备注": ""
        })
        testcase_id += 1

    return testcases


def generate_api_testcases(api_info: dict, module_name: str):
    """生成接口测试用例"""
    return generate_api_testcases(api_info, module_name)


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        print("Usage: python testcase_generator.py <type> <module_name> [data_json]")
        print("  type: functional | api")
        print("  module_name: 模块名称")
        print("  data_json: 可选，JSON格式的数据")
        sys.exit(1)

    type_arg = sys.argv[1]
    module = sys.argv[2]
    data_json = sys.argv[3] if len(sys.argv) > 3 else "{}"

    try:
        data = json.loads(data_json)
    except json.JSONDecodeError:
        data = {}

    if type_arg == "functional":
        testcases = generate_functional_testcases({"name": module, "description": ""}, module)
        filename = create_functional_testcase_excel(testcases, module)
    elif type_arg == "api":
        testcases = generate_api_testcases(data, module)
        filename = create_api_testcase_excel(testcases, module)
    else:
        print(f"Unknown type: {type_arg}")
        sys.exit(1)

    print(f"Generated: {filename}")
